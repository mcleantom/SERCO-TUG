# -*- coding: utf-8 -*-
"""
Created on Thu Apr  8 10:10:44 2021

@author: Rastko
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import datetime
from scipy.interpolate import interp1d
from openpyxl.styles import Alignment
import calendar

# =============================================================================
# INPUTS
# =============================================================================

engine_curves = pd.read_excel("Data/engine_curve.xlsx")
csv_files = ["Data\engine_data_2021-02.csv",
             "Data\engine_data_2021-03.csv"]

# =============================================================================
# CODE
# =============================================================================

def read_csv(filepaths):
    
    dfs = []
    
    for filepath in filepaths:
        df = pd.read_csv(filepath, usecols=["imei", "dataloggertime", "longitude",
                                            "latitude", "speedoverground",
                                            "revolutions", "enginetorque",
                                            "fuelrate", "bus_id", "percentload"])
        dfs.append(df)
    
    df = pd.concat(dfs)
    df.dropna(subset = ["imei"])
    df["dataloggertime"] = pd.to_datetime(df["dataloggertime"])
    df = df.sort_values(by=["dataloggertime"])
    df = df.set_index(["dataloggertime"])
    
    engines = split_engines(df)
    for i in engines:
        engines[i] = calc_torque(engines[i])
        engines[i] = calc_power(engines[i])
    
    return df, engines


def split_engines(df):
    return dict(tuple(df.groupby('bus_id')))


def split_days(df):
    return dict(tuple(df.groupby([df.index.year, df.index.month, df.index.day])))


def split_seconds(df):
    return df.groupby([df.index.year,
                       df.index.month,
                       df.index.day,
                       df.index.minute,
                       df.index.second])


def calc_torque(df):
    """
    Fit the engine RPM to a torque curve
    """
    # df["torque"] = np.nan

    # mask = df["revolutions"] < 1025

    # rpm = df.loc[mask, "revolutions"]
    # df.loc[mask, "torque"] = (0.00001698*rpm**3 -
    #                           0.0273*rpm**2 +
    #                           14.99*rpm + 1276)

    # mask = df["revolutions"] >= 1025

    # rpm = df.loc[mask, "revolutions"]
    # df.loc[mask, "torque"] = (-0.0000005388274*rpm**4 +
    #                           +0.003072432*rpm**3 -
    #                           6.551486*rpm**2 +
    #                           6186.143*rpm - 2171670)
    x = [1600,1500,1400,1300,1200,1150,1100,1000,900,800,650,400]
    y = [7789,8308,8901,9400,9486,8400,7084,5854,5305,4381,4114,4000]
    f = interp1d(x, y, kind="linear", bounds_error=False, fill_value="extrapolate")
    df["torque"] = f(df["revolutions"])
    # df["torque"] = 
    
    return df


def calc_power(df):
    df["power"] = (df["percentload"]/100)*2*np.pi*(df["revolutions"]/60)*(df["torque"]/1000)
    return df


def sum_engine_powers(engines):
    combined = engines[1].join(engines[2], how="outer", rsuffix="_2")
    combined["total_power"] = combined["power"] + combined["power_2"]
    combined["ave_rpm"] = (combined["revolutions"] + combined["revolutions_2"])/2
    combined["speedoverground"] = combined["speedoverground"].replace(to_replace=0, method='ffill')
    combined["total_fuel_rate"] = combined["fuelrate"] + combined["fuelrate_2"]
    return combined


def plot_power_vs_rpm(df, title=""):
    fig, ax = plt.subplots(dpi=512, figsize=(6.51, 3.93)) 
    ax.plot(df["ave_rpm"], df["total_power"], 'o', markersize=0.5, color="blue")
    cols = engine_curves.columns[1:]
    for col in cols:
        plt.plot(engine_curves["RPM"], engine_curves[col], label="Engine Power "+str(col)+"%")
   

    ax.set_xlim(left=600)#(600, 1600)
    ax.set_ylim(bottom=0)#0, 2500)
    ax.set_xlabel("Engine RPM")
    ax.set_ylabel("Total Power [kW]")
    ax.legend(loc="upper left")
    ax.grid()
    fig.tight_layout()
    if title:
        title += "_"
    fig.savefig("Figures/"+title+"power_vs_rpm.png")


def plot_power_vs_sog(df, title=""):
    fig, ax = plt.subplots(dpi=512, figsize=(6.51, 3.93))
    
    ax.plot(df["speedoverground"], df["total_power"], 'o', markersize=0.5, color="blue")

    cols = engine_curves.columns[1:]
    for col in cols:
        ax.plot([0, 7], [engine_curves[col][0]]*2, label="Engine Power "+str(col)+"%")

    ax.set_xlim(left=0)#(0, 7)
    ax.set_ylim(bottom=0)#(0, 2500)
    ax.set_xlabel("SOG [kts]")
    ax.set_ylabel("Total Power [kW]")
    ax.legend(loc="upper left")
    ax.grid()
    fig.tight_layout()
    if title:
        title += "_"
    fig.savefig("Figures/"+title+"power_vs_sog.png")


def plot_day(df):
    
    start = df.index[0].to_pydatetime().date()
    end = start# + datetime.timedelta(days=1)
    start = datetime.datetime(start.year, start.month, start.day, 6)
    end = datetime.datetime(end.year, end.month, end.day, 20)
    fig, ax1 = plt.subplots(dpi=512, figsize=(6.48, 2.95))
    
    ax1.set_ylabel("Total Power [kW], RPM")
    df["total_power"].plot(ax=ax1, x_compat=True, color="black", label="Total Power", linewidth=0.8)
    df["ave_rpm"].plot(ax=ax1, x_compat=True, color="blue", label="RPM", linewidth=0.8)

    ax1.set_xlim([start, end])
    ax1.grid()
    ax1.legend(loc="upper left")

    xticks = pd.date_range(start, end, freq='H')
    ax1.set_xticklabels([x.strftime('%H') for x in xticks])
    ax1.xaxis.set_tick_params(rotation=00)

    ax2 = ax1.twinx()
    ax2.set_ylabel('SOG [knots]')
    
    df[df["speedoverground"]>0]["speedoverground"].plot(ax=ax2, x_compat=True, color="red", label="SOG", linewidth=0.8)
    
    ax2.legend(loc="upper right")
    
    ax1.xaxis.set_major_locator(mdates.HourLocator(interval=1))
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H"))
    
    ax1.set_xlabel("Hour of the day [hrs]")
    ax1.set_ylim(left=0)#(0, 2500])
    ax2.set_ylim(left=0)#[0, 10])


def split_trips(df):
    g = (df.index.to_series().diff().dt.seconds > 1).cumsum()
    trips = dict(tuple(df.groupby(g)))
    for trip in trips:
        trips[trip] = trips[trip][trips[trip]["imei"].notna()]
        trips[trip] = trips[trip][trips[trip]["imei_2"].notna()]
    return trips


def plot_trips(df):
    
    start = df.index[0].to_pydatetime().date()
    end = start# + datetime.timedelta(days=1)
    start = datetime.datetime(start.year, start.month, start.day, 6)
    end = datetime.datetime(end.year, end.month, end.day, 20)
    
    fig, ax1 = plt.subplots(dpi=512)
    ax1.set_ylabel("Total Power [kW], RPM")
    ax1.set_xlim([start, end])
    
    trips = split_trips(df)
    for trip in trips:
        trips[trip]["ave_rpm"].plot(ax=ax1, x_compat=True, label="Trip " + str(trip))

    xticks = pd.date_range(start, end, freq='H')
    ax1.set_xticklabels([x.strftime('%H') for x in xticks])
    ax1.xaxis.set_tick_params(rotation=00)
    ax1.xaxis.set_major_locator(mdates.HourLocator(interval=1))
    ax1.xaxis.set_major_formatter(mdates.DateFormatter("%H"))
    
    ax1.grid()
    ax1.set_xlabel("Hour of the day [hrs]")
    ax1.set_ylim(bottom=0)#[0, 2500])
    ax1.legend()


def valid_trip(df):
    flag = True
    if len(df) == 0:
        flag = False
    if df["total_power"].sum()/3600 < 5:
        flag = False
    return flag

def analyse_trips(df):
    
    results = {}

    for i in df:
        trip = df[i]
        
        if valid_trip(trip):
        # if len(trip) > 0:
            mask = trip["total_power"]>=5
            trip_result = {'start_time': trip.index[0],
                           'stop_time': trip.index[-1],
                           'duration': (trip.index[-1] - trip.index[0])/datetime.timedelta(hours=1),
                           'energy_used': (trip["total_power"].sum())/3600,
                           'total_fuel': (trip["total_fuel_rate"]/3600).sum(),
                           'fuel_at_power': (trip[mask]["total_fuel_rate"]/3600).sum(),
                           'fuel_idle': (trip[~mask]["total_fuel_rate"]/3600).sum(),
                           'duration_at_power': len(trip[mask])/3600,
                           'duration_idle': len(trip[~mask])/3600
                           }

            results["Trip " + str(i)] = trip_result
            
    results = pd.DataFrame.from_dict(results).T
    results["time_to_charge"] = (results["start_time"] - results["stop_time"].shift(1)).fillna(pd.Timedelta(seconds=0))
    # charge_time = results["time_to_charge"]/datetime.timedelta(hours=1)
    # days, hours = divmod(charge_time, 24)
    # results["altered_time_to_charge"] = 24*(days>0) + hours
    results["time_to_charge"] = calc_time_to_charge(results)
    results["first_trip"] = np.invert(results["start_time"].dt.day == results["start_time"].shift(1).dt.day)
    results["date"] = results["start_time"].dt.strftime('%d-%m')
    # results["total_fuel_used"] = (results["total_fuel_rate"]/3600).sum()
    # mask = results["total_power"]<=5
    # results["fuel_used_at_power"] = (results[mask]["total_fuel_rate"]/3600).sum()
    cols = ['duration', 'energy_used', 'total_fuel',
       'fuel_at_power', 'fuel_idle', 'duration_at_power', 'duration_idle',
       'time_to_charge']
    results[cols] = results[cols].astype(float)
    return results


def analyse_days(df):
    # days = dict(tuple(df.groupby('date')))
    days = df.groupby('date')
    results = {}
    for i, day in days:
        day_result = {'total_charge_time': day["time_to_charge"].sum(),
                      'total_energy': day["energy_used"].sum(),
                      'num_trips':len(day)}
        results[day["start_time"].dt.date[0]] = day_result
    
    results = pd.DataFrame.from_dict(results).T
    results = results.sort_index()
    results["cum_num_trips"] = results["num_trips"].cumsum()+1

    return results
    

def plot_charge_time_vs_energy(df):
    
    fig, ax1 = plt.subplots(dpi=512, figsize=(6.48, 2.95))

    charge_time = df["time_to_charge"]#/datetime.timedelta(hours=1)
    # days, hours = divmod(charge_time, 24)
    # altered_times = -24*(days>0) + hours

    mask = df["first_trip"]
    ax1.plot(df[mask]["time_to_charge"], df[mask]["energy_used"], '^', color="blue", label="First trip of the day")
    ax1.plot(df[~mask]["time_to_charge"], df[~mask]["energy_used"], 'o', color="red", label="Other trips")
    
    # ax1.plot(altered_times, df["energy_used"], 'o', color="blue")
    ax1.set_xlabel("Time to Charge per day [hours]")
    ax1.set_ylabel("Energy Used per Trip [kWh]")
    ax1.set_xlim(left=0)#[0, 32])
    ax1.set_ylim(bottom=0)#[0, 800])
    ax1.grid()
    ax1.legend(loc="upper right")
    fig.tight_layout()
    fig.savefig("energy_used_vs_charge_time (trips).svg")


def plot_energy_used_vs_time_to_charge(df):
    
    fig, ax1 = plt.subplots(dpi=512, figsize=(6.48, 2.95))
    mask = df["total_charge_time"] != 0
    ax1.plot(df[mask]["total_charge_time"], df[mask]["total_energy"], "o", color="blue")
    ax1.set_xlim(left=0)#[0, 32])
    ax1.set_ylim(bottom=0)#[0, 1200])
    ax1.set_xlabel("Time to Charge per day [hours]")
    ax1.set_ylabel("Energy Used per day [kWh]")
    ax1.grid()
    fig.tight_layout()
    fig.savefig("energy_used_vs_charge_time (days).svg")


def formatted_table(df):
    export = pd.DataFrame(df["date"])
    export["Duration [hr]"] = df["duration"]
    export["Time to charge [hr]"] = df["time_to_charge"]
    export["Energy used [kWh]"] = df["energy_used"]
    export["Fuel used at power [litre]"] = df["fuel_at_power"]
    export["Total Fuel used [liter]"] = df["total_fuel"]
    return export


def calc_time_to_charge(df):
    
    start = df["start_time"]
    stop = df["stop_time"]
    
    day_start_df = pd.DataFrame({'year': start.dt.year,
                            'month': start.dt.month,
                            'day': start.dt.day})
    day_start_df = pd.to_datetime(day_start_df, utc=True)
    time_start_df = start - day_start_df

    day_stop_df = pd.DataFrame({'year': stop.dt.year,
                            'month': stop.dt.month,
                            'day': stop.dt.day})
    day_stop_df = pd.to_datetime(day_stop_df, utc=True)
    time_stop_df = stop - day_stop_df
    
    new_day = (start.shift(1).dt.date != stop.dt.date)
    
    day_shift = new_day.cumsum() * datetime.timedelta(days=1)
    
    time_start_df += day_shift
    time_stop_df += day_shift
    
    time_to_charge = (time_start_df - time_stop_df.shift(1))/datetime.timedelta(hours=1)

    return time_to_charge


def col_num_to_string(n): return 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[n] if n <= 25  else col_num_to_string(divmod(n,26)[0]-1) + 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'[divmod(n,26)[1]]


def string_to_col_num(string):
    return sum([(26**i)*(ord(c)-64) for i, c in enumerate(string[::-1].upper())])


def save_to_excel(df, day_df, month_df, month_hist_data):
    date_cols = df.select_dtypes(include=['datetime64[ns, UTC]']).columns
    for date_col in date_cols:
        df[date_col] = df[date_col].dt.tz_localize(None)
    
    formatted_df = formatted_table(df)
    path = "results.xlsx"
    last_row = 1

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # workbook = writer.book
        df.to_excel(writer, sheet_name="raw_data")
        month_df.to_excel(writer, sheet_name="Monthly Data")
        month_hist_data.to_excel(writer, sheet_name="Histogram")
        formatted_df.to_excel(writer, sheet_name="formatted_table")
        worksheet = writer.sheets["formatted_table"]
        for row in day_df.itertuples():
            # left_cell = worksheet.cell(column=8, row=last_row+1)
            # right_cell = worksheet.cell(column=9, row=last_row+1)
            # left_cell.value = row[1]
            # right_cell.value = row[2]
            worksheet.cell(column=8, row=last_row+1, value=row[1])
            worksheet.cell(column=9, row=last_row+1, value=row[2])
            # worksheet.cell(row=row[4], column=string_to_col_num("H")).value = row[1]
            # worksheet.cell(row=row[4], column=string_to_col_num("I")).value = row[2]
            # worksheet.merge_range(last_row+1, row[4], 8, 8, row[1])
            # worksheet.merge_range(last_row+1, row[4], 9, 9, row[2])
            worksheet.merge_cells(start_row=last_row+1, start_column=2, end_row=row[4], end_column=2)
            worksheet.merge_cells(start_row=last_row+1, start_column=8, end_row = row[4], end_column = 8)
            worksheet.merge_cells(start_row=last_row+1, start_column=9, end_row=row[4], end_column=9)
            
            last_row = row[4]
        for col in 'ABCDEFJHI':
            active_col = worksheet.column_dimensions[col]
            active_col.alignment = Alignment(horizontal="center", vertical="center")
        

def plot_days(df):
    df = df.copy()
    wrk = df.resample('D', on='start_time').sum()
    dayNo = wrk.index.size
    print(dayNo)
    
    start_date = wrk.index.date.min().isocalendar()
    start_week = start_date[1]
    start_year = start_date[0]
    d = str(start_year) +"-W"+str(start_week)
    resample_start = datetime.datetime.strptime(d + "-1", "%Y-W%W-%w")
    wrk = wrk.reindex(pd.date_range(start = resample_start,
                                    periods = dayNo - (dayNo % -7)), fill_value=0)

    wrk["week"] = wrk.index.isocalendar().week
    wrk.index = wrk.index.astype(str)
    for week, grp in wrk.groupby('week'):
        fig, ax = plt.subplots(figsize=(6.48, 2.95), dpi=512)
        grp.iloc[:, :2].plot.bar(ax=ax, stacked=True, color=["tab:orange", "tab:grey"])

def plot_weeks(df):
    df = df.copy()
    wrk = df.resample('W', on='start_time').sum()
    dayNo = wrk.index.size
    print(dayNo)
    
    start_date = wrk.index.date.min().isocalendar()
    start_week = start_date[1]
    start_year = start_date[0]
    d = str(start_year) +"-W"+str(start_week)
    resample_start = datetime.datetime.strptime(d + "-1", "%Y-W%W-%w")
    wrk.index = wrk.index - datetime.timedelta(days=6)
    num_weeks = (wrk.index.max()-wrk.index.min()).days/7 + 1
    wrk = wrk.reindex(pd.date_range(start = resample_start,
                                    periods = num_weeks,
                                    freq="W")-datetime.timedelta(days=6), fill_value=0)

    # wrk["week"] = wrk.index.isocalendar().week
    # wrk.index = wrk.index.astype(str)
    
    fig, ax = plt.subplots(figsize=(6.48, 2.95), dpi=512)
    wrk.plot.bar(ax=ax, stacked=True, color=["tab:orange", "tab:grey"])
    ax.set_xticklabels(list(wrk.index.strftime("%Y-%b")))


def analyse_months(df):
    cols = ["start_time", 'duration', 'energy_used', 'total_fuel',
       'fuel_at_power', 'fuel_idle', 'duration_at_power', 'duration_idle',
       'time_to_charge']
    df = df[cols].copy()
    wrk = df.resample('M', on='start_time').sum()
    start_date = wrk.index.date.min().isocalendar()
    start_week = start_date[1]
    start_year = start_date[0]
    d = str(start_year) +"-W"+str(start_week)
    wrk.index = [x.replace(day=1) for x in wrk.index]
    date_range = pd.date_range(start=wrk.index.min(), periods=6, freq="M")
    date_range = [x.replace(day=1) for x in date_range]
    wrk = wrk.reindex(date_range, fill_value=np.nan)
    hours_in_month = np.array([calendar.monthrange(x.year, x.month)[1]*24 for x in wrk.index])
    wrk["duration_engine_off"] = hours_in_month - wrk["duration"]
    wrk["total_co2"] = wrk["total_fuel"] * 2.675994109
    wrk = wrk[["duration_engine_off", "duration", "duration_idle",
               "duration_at_power", "fuel_idle", "fuel_at_power", "total_fuel",
               "total_co2"]]
    wrk.index = wrk.index.date
    # wrk = wrk.T
    wrk.loc["total"] = wrk.sum(axis=0)
    return wrk
    

def plot_months(df):
    df = df.copy()
    wrk = df.resample('M', on='start_time').sum()
    dayNo = wrk.index.size
    print(dayNo)
    
    start_date = wrk.index.date.min().isocalendar()
    start_week = start_date[1]
    start_year = start_date[0]
    d = str(start_year) +"-W"+str(start_week)
    resample_start = datetime.datetime.strptime(d + "-1", "%Y-W%W-%w")
    wrk.index = [x.replace(day=1) for x in wrk.index]
    date_range = pd.date_range(start=wrk.index.min(), periods=6, freq="M")
    date_range = [x.replace(day=1) for x in date_range]
    wrk = wrk.reindex(date_range, fill_value=0)
    
    fig, ax = plt.subplots(figsize=(6.48, 2.95), dpi=512)
    wrk.plot.bar(ax=ax, stacked=True, color=["tab:orange", "tab:grey"])
    ax.set_xticklabels(wrk.index.strftime("%Y-%b"))
        

    # return wrk
    # min_week = wrk.index.date.min().isocalendar()[1]
    
    # wrk = 
    # weeks = df.groupby(df["start_time"].dt.week)
    # for _, week in weeks:
    #     week.groupby(week["start_time"].dt.day)[["duration_idle", "duration_at_power"]].sum().plot.bar(stacked=True)
    # return week


def calc_histogram_engine_data(df, title="count"):
    n, bins = np.histogram(df["total_power"], bins=[5,100,200,300,400,500,600,700,800,900,1000,1100,1200,1300,1400,1500,1600,1700,1800,1900,2000,2100,2200,2300,2400])
    width = np.diff(bins)
    bins = [str(x) for x in bins]
    bins = [x[0]+"-"+x[1] for x in zip(bins[:-1], bins[1:])]
    n = n/60        
    return pd.DataFrame(n, index=bins, columns=[title])


def plot_histogram_engine_data(df):
    n, bins = np.histogram(df["total_power"], bins=[5,100,200,300,400,500,600,700,800,900,1000,1100,1200,1300,1400,1500,1600,1700,1800,1900,2000,2100,2200,2300,2400])
    width = np.diff(bins)
    bins = [str(x) for x in bins]
    bins = [x[0]+"-"+x[1] for x in zip(bins[:-1], bins[1:])]
    n = n/60
    fig, ax = plt.subplots(dpi=512, figsize=(6.51, 3.93))
    # ax.grid(True, which="minor")
    # ax.grid(which="both", zorder=0)
    # ax.xaxis.set_minor_locator(AutoMinorLocator())
    rects = ax.bar(bins, n, zorder=3)
    ax.tick_params(axis='x', rotation=90)
    # ax.set_ylim(0, 900)
    for rect in rects:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., height+10,
                '%d' % int(height),
                ha='center', va='bottom', fontsize="x-small")


def monthly_hist_data(df):
    hist_data = []
    months = df.groupby(df.index.month)
    for i, month in months:
        hist_data += [calc_histogram_engine_data(month,
                                                 title=month.index[0].strftime("%Y-%m"))]
    
    temp = hist_data[0]
    for x in hist_data[1:]:
        temp = temp.join(x)
    temp["total"] = temp.sum(axis=1)
    return temp


def plot_monthly_power_vs_rpm_and_sog(df):
    months = df.groupby(df.index.month)
    for i, month in months:
        plot_power_vs_rpm(month, title=month.index[0].strftime("%Y-%m"))
        plot_power_vs_sog(month, title=month.index[0].strftime("%Y-%m"))

# file = "Data\engine_data_2021-02.csv"
# data_1, engines_1 = read_csv(file)
# file = "Data\engine_data_2021-03.csv"
# data_2, engines_2 = read_csv(file)

# data = pd.concat([data_1, data_2])
# engines = pd.concat([engines_1, engines_2])

data, engines = read_csv(csv_files)

combined = sum_engine_powers(engines)
plot_power_vs_rpm(combined, title="all_data")
plot_power_vs_sog(combined, title="all_data")
plot_monthly_power_vs_rpm_and_sog(combined)
days = split_days(combined)

# histogram_data = calc_histogram_engine_data()
trips = split_trips(combined)
trip_results = analyse_trips(trips)
plot_charge_time_vs_energy(trip_results)
day_results = analyse_days(trip_results)
month_results = analyse_months(trip_results)
month_hist_data = monthly_hist_data(combined)
plot_energy_used_vs_time_to_charge(day_results)
save_to_excel(trip_results, day_results, month_results.T, month_hist_data)
plot_months(trip_results[["start_time", "duration_idle", "duration_at_power"]])
plot_months(trip_results[["start_time", "fuel_idle", "fuel_at_power"]])